"""
POS sales importer for URS Majestic.

Steps:
  1. Upserts menu_items from backend/data/menu_seed.json (canonical source of truth).
  2. Parses backend/data/pos_export.xlsx (Item | Date | Qty | Total columns);
     drops any rows dated today (same-day exports are incomplete).
  3. Resolves each POS name to a canonical menu item via pos_name_map then exact match.
  4. Replaces item_sales for the date range covered by the file (not the
     whole table — a corrected re-export heals only its own dates).
  5. Upserts one daily_channel_sales row per date (channel='petpooja').
  6. Prints a summary.

Usage (from backend/ with venv active):
    python -m scripts.import_pos
    python -m scripts.import_pos --xlsx path/to/other.xlsx
"""

import sys
import os
import argparse
import json
from datetime import date
from decimal import Decimal

# Windows consoles default to cp1252, which can't encode the rupee sign this
# script prints in its summary -- crashes AFTER the DB commit, making a
# successful import look like a failure. Linux (Render) is already UTF-8.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

# Allow running from project root or backend/
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".env")
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip().strip('"').strip("'"))

from collections import defaultdict

from sqlalchemy import func
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core.database import SessionLocal
from app.models.menu_item import MenuItem, PosAlias
from app.models.item_sale import ItemSale
from app.models.daily_channel_sales import DailyChannelSales
from app.models.upload_log import UploadLog
from app.services.sales_stock import adjust_stock_for_sales
from app.services.order_derived_stock import sync_order_derived_stock

SEED_JSON = os.path.join(os.path.dirname(__file__), "..", "data", "menu_seed.json")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "..", "data", "pos_export.xlsx")


# -- Step 1: Seed menu items ------------------------------------------─────────────────

def seed_menu_items(db, seed: dict) -> dict[str, MenuItem]:
    """Upsert menu items from seed data. Returns name→MenuItem map."""
    food_cost_default = Decimal(str(seed["_meta"]["food_cost_pct_default"]))
    existing: dict[str, MenuItem] = {m.name: m for m in db.query(MenuItem).all()}
    inserted = updated = 0

    for row in seed["items"]:
        name = row["name"]
        price = Decimal(str(row["price"]))
        if name in existing:
            item = existing[name]
            changed = False
            for field, val in [("category", row["category"]), ("is_food", row["is_food"])]:
                if getattr(item, field) != val:
                    setattr(item, field, val)
                    changed = True
            if item.price != price:
                item.price = price
                changed = True
            if item.food_cost_pct != food_cost_default:
                item.food_cost_pct = food_cost_default
                changed = True
            if changed:
                updated += 1
        else:
            item = MenuItem(
                name=name,
                category=row["category"],
                price=price,
                is_active=True,
                is_food=row["is_food"],
                food_cost_pct=food_cost_default,
            )
            db.add(item)
            db.flush()
            existing[name] = item
            inserted += 1

    # Upsert POS aliases
    existing_aliases: set[str] = {a.pos_name for a in db.query(PosAlias).all()}
    alias_added = 0
    for pos_name, canonical in seed.get("pos_name_map", {}).items():
        if pos_name not in existing_aliases:
            item = existing.get(canonical)
            if item:
                db.add(PosAlias(menu_item_id=item.id, pos_name=pos_name))
                alias_added += 1

    db.flush()
    print(f"  [menu_items]  {inserted} inserted, {updated} updated")
    print(f"  [pos_aliases] {alias_added} added")
    return existing


class ParseError:
    def __init__(self, line: int, message: str):
        self.line = line
        self.message = message

    def __repr__(self):
        return f"ParseError(line={self.line}, message={self.message!r})"


# ── Step 2: Parse xlsx ────────────────────────────────────────────────────────

def parse_xlsx(path: str) -> tuple[list[dict], list[ParseError], Decimal | None, int | None]:
    """Parse POS xlsx. Returns (rows, parse_errors, file_declared_total,
    file_declared_rows) where rows are {raw_name, sale_date, qty, revenue}
    and the declared total/rows come from the file's own 'Total' summary row
    (the file's self-reported checksum — used for upload_log auditing, never
    for the actual insert)."""
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find header row: look for a row where col A contains "Item" (case-insensitive)
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).strip().lower() == "item":
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find header row with 'Item' in column A")

    rows: list[dict] = []
    errors: list[ParseError] = []
    declared_total: Decimal | None = None
    declared_rows: int | None = None

    for line_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw_name = row[0]
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()
        if not raw_name:
            continue

        date_val, qty_val, revenue_val = row[1], row[2], row[3]

        if raw_name.lower() in ("total", "grand total"):
            # The file's own self-reported checksum row — not a data row.
            try:
                if qty_val is not None:
                    declared_rows = int(qty_val)
                if revenue_val is not None:
                    declared_total = Decimal(str(revenue_val))
            except Exception as exc:
                errors.append(ParseError(line_no, f"Could not read declared total row: {exc}"))
            continue

        if date_val is None or qty_val is None or revenue_val is None:
            errors.append(ParseError(line_no, f"Row for {raw_name!r} missing date/qty/total"))
            continue

        # Normalize date
        if isinstance(date_val, (date,)):
            sale_date = date_val
        else:
            from datetime import datetime as dt
            if hasattr(date_val, "date"):
                sale_date = date_val.date()
            else:
                try:
                    sale_date = dt.strptime(str(date_val).strip(), "%d-%m-%Y").date()
                except ValueError:
                    try:
                        sale_date = dt.strptime(str(date_val).strip(), "%Y-%m-%d").date()
                    except ValueError:
                        errors.append(ParseError(line_no, f"Could not parse date {date_val!r} for item {raw_name!r}"))
                        continue

        try:
            qty = Decimal(str(qty_val))
            revenue = Decimal(str(revenue_val))
        except Exception as exc:
            errors.append(ParseError(line_no, f"Non-numeric qty/total for {raw_name!r}: {exc}"))
            continue

        rows.append({
            "raw_name": raw_name,
            "sale_date": sale_date,
            "qty": qty,
            "revenue": revenue,
        })

    return rows, errors, declared_total, declared_rows


# ── Step 2b: Drop same-day rows ────────────────────────────────────────────────

def exclude_today(rows: list[dict], today) -> tuple[list[dict], int]:
    """Same-day exports are incomplete by definition — drop rows dated today.
    Returns (kept_rows, excluded_count)."""
    kept = [r for r in rows if r["sale_date"] != today]
    return kept, len(rows) - len(kept)


# ── Step 3: Resolve names ─────────────────────────────────────────────────────

def build_resolver(seed: dict, menu_map: dict[str, MenuItem]) -> dict[str, str]:
    """Returns pos_name → canonical_name mapping for all known names."""
    resolver: dict[str, str] = {}
    # From pos_name_map
    for pos_name, canonical in seed.get("pos_name_map", {}).items():
        resolver[pos_name] = canonical
    # Direct exact matches
    for name in menu_map:
        if name not in resolver:
            resolver[name] = name
    return resolver


# ── Step 4: Insert item_sales ─────────────────────────────────────────────────

def load_sales(db, rows: list[dict], resolver: dict[str, str]) -> tuple[list, list]:
    """Replace item_sales for the date range covered by `rows` (not the whole
    table) so a corrected re-export heals just its own dates without wiping
    history outside that range.

    Every row is inserted — never silently dropped. A row with no
    pos_aliases/menu_items match still gets a row, with item_name falling
    back to raw_name, so revenue always reconciles against
    daily_channel_sales (see v_recon_daily). `unmatched` is still reported
    separately so unmapped names stay visible."""
    to_insert = []
    unmatched = []

    for row in rows:
        canonical = resolver.get(row["raw_name"])
        if canonical is None:
            unmatched.append(row["raw_name"])
        to_insert.append(ItemSale(
            raw_name=row["raw_name"],
            item_name=canonical or row["raw_name"],
            sale_date=row["sale_date"],
            qty=row["qty"],
            revenue=row["revenue"],
        ))

    if rows:
        min_date = min(r["sale_date"] for r in rows)
        max_date = max(r["sale_date"] for r in rows)
        db.query(ItemSale).filter(
            ItemSale.sale_date >= min_date, ItemSale.sale_date <= max_date
        ).delete(synchronize_session=False)

    db.bulk_save_objects(to_insert)
    db.flush()

    return to_insert, unmatched


# ── Step 5: Upsert daily_channel_sales ──────────────────────────────────────────

def upsert_daily_channel_sales(db, rows: list[dict], source_file: str) -> int:
    """Aggregate parsed rows (all of them — resolved or not; totals are
    pre-tax with no discounts in Petpooja) into one net_sales figure per
    business_date, upserted as channel='petpooja'. Returns dates touched."""
    totals: dict = defaultdict(Decimal)
    for row in rows:
        totals[row["sale_date"]] += row["revenue"]

    for sale_date, net_sales in totals.items():
        stmt = pg_insert(DailyChannelSales).values(
            business_date=sale_date,
            channel="petpooja",
            net_sales=net_sales,
            orders=None,
            gross_order_value=None,
            restaurant_discount=Decimal("0"),
            platform_discount=Decimal("0"),
            source_file=source_file,
        )
        stmt = stmt.on_conflict_do_update(
            index_elements=["business_date", "channel"],
            set_={
                "net_sales": stmt.excluded.net_sales,
                "source_file": stmt.excluded.source_file,
                "uploaded_at": func.now(),
            },
        )
        db.execute(stmt)

    db.flush()
    return len(totals)


# ── Step 6: Log the upload ──────────────────────────────────────────────────────

def write_upload_log(
    db,
    *,
    source_file: str,
    rows: list[dict],
    parse_errors: list[ParseError],
    excluded_today: int,
    declared_total: Decimal | None,
    declared_rows: int | None,
    succeeded: bool,
    error_detail: str | None = None,
) -> UploadLog:
    """Every upload writes exactly one upload_log row, success or failure —
    silent row-dropping is forbidden, so this is the audit trail proving it."""
    dates = [r["sale_date"] for r in rows]
    log = UploadLog(
        channel="petpooja",
        source_file=source_file,
        period_start=min(dates) if dates else None,
        period_end=max(dates) if dates else None,
        file_declared_total=declared_total,
        file_declared_rows=declared_rows,
        rows_parsed=len(rows) + excluded_today + len(parse_errors),
        rows_inserted=len(rows),
        rows_skipped_today=excluded_today,
        rows_failed=len(parse_errors),
        amount_inserted=sum((r["revenue"] for r in rows), Decimal("0")) if rows else Decimal("0"),
        succeeded=succeeded,
        error_detail=error_detail,
    )
    db.add(log)
    db.flush()
    return log


# ── Main ──────────────────────────────────────────────────────────────────────

def main(xlsx_path: str):
    with open(SEED_JSON) as f:
        seed = json.load(f)

    db = SessionLocal()
    try:
        print("\n-- Step 1: Seed menu items ------------------------------------------")
        menu_map = seed_menu_items(db, seed)

        print("\n-- Step 2: Parse POS xlsx -------------------------------------------")
        raw_rows, parse_errors, declared_total, declared_rows = parse_xlsx(xlsx_path)
        raw_rows, excluded_today = exclude_today(raw_rows, date.today())
        print(f"  {len(raw_rows)} data rows read from {os.path.basename(xlsx_path)}")
        if excluded_today:
            print(f"  {excluded_today} row(s) dated today excluded (incomplete same-day export)")
        if parse_errors:
            print(f"  {len(parse_errors)} row(s) failed to parse:")
            for e in parse_errors:
                print(f"    line {e.line}: {e.message}")

        print("\n-- Step 3: Resolve and insert item_sales ----------------------------")
        resolver = build_resolver(seed, menu_map)
        matched, unmatched = load_sales(db, raw_rows, resolver)

        print("\n-- Step 4: Calculate ingredient usage and adjust Stock Log ----------")
        stock_result = adjust_stock_for_sales(db, matched)
        print(f"  {stock_result['adjusted']} stock balance(s) adjusted")
        print(f"  {stock_result['unchanged']} unchanged usage balance(s) skipped")
        if stock_result["initialized_zero"]:
            print(f"  {stock_result['initialized_zero']} ingredient(s) initialized at zero: no compatible prior count")
        if stock_result["unresolved"]:
            print(f"  {len(stock_result['unresolved'])} recipe input(s) could not be calculated")

        try:
            with db.begin_nested():
                derived_result = sync_order_derived_stock(db)
            print(f"  order-derived model: {derived_result['deductions_applied']} deduction(s), "
                  f"{derived_result['additions_applied']} purchase addition(s) applied "
                  f"({derived_result['skipped_no_baseline']} skipped, no baseline yet)")
        except Exception as exc:
            print(f"  order-derived model pass failed (experimental, non-blocking): {exc}")

        print("\n-- Step 5: Upsert daily_channel_sales --------------------------------")
        days = upsert_daily_channel_sales(db, raw_rows, os.path.basename(xlsx_path))
        print(f"  {days} business date(s) upserted (channel=petpooja)")

        print("\n-- Step 6: Log upload -------------------------------------------------")
        write_upload_log(
            db,
            source_file=os.path.basename(xlsx_path),
            rows=raw_rows,
            parse_errors=parse_errors,
            excluded_today=excluded_today,
            declared_total=declared_total,
            declared_rows=declared_rows,
            succeeded=True,
        )

        db.commit()

        # ── Summary ───────────────────────────────────────────────────────────
        dates = [r.sale_date for r in matched]
        total_rev = sum(r.revenue for r in matched)
        distinct_items = len({r.item_name for r in matched})
        unmatched_unique = sorted(set(unmatched))

        print("\n==============================================================")
        print("  IMPORT SUMMARY")
        print("==============================================================")
        print(f"  Rows matched   : {len(matched)}")
        print(f"  Rows unmatched : {len(unmatched)}")
        print(f"  Distinct items : {distinct_items}")
        print(f"  Total revenue  : ₹{total_rev:,.2f}")
        if dates:
            print(f"  Date range     : {min(dates)} → {max(dates)}")
        if unmatched_unique:
            print(f"\n  Unmatched POS names ({len(unmatched_unique)}):")
            for n in unmatched_unique:
                print(f"    • {n}")
        print("==============================================================\n")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import POS sales xlsx into ursmajestic DB")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to POS xlsx export")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: xlsx not found at {args.xlsx}")
        print("Place your POS export at backend/data/pos_export.xlsx or pass --xlsx <path>")
        sys.exit(1)

    main(args.xlsx)
