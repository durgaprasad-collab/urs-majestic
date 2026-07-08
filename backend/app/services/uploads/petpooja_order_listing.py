"""Parses Petpooja's 'Order Listing' xlsx (Reports → Order Report).

Recovers per-day COUNTER order counts (and cross-checks their total against
the item-report-derived net_sales) — it never touches net_sales itself,
which stays owned by the item-report upload (see scripts/import_pos.py).

Order No. resets daily and repeats across days, so it is not a usable
idempotency key; idempotency instead comes from recomputing per-day counts
from the file each time and overwriting daily_channel_sales.orders."""
import decimal
from dataclasses import dataclass
from datetime import date, datetime, timedelta

from sqlalchemy.orm import Session

from app.models.daily_channel_sales import DailyChannelSales

REQUIRED_COLUMNS = ["Order No.", "Order Type", "My Amount (₹)", "Total Discount (₹)", "Status", "Created"]

ALARM_CUTOFF = date(2026, 6, 29)
ALARM_MESSAGE = "Aggregator orders are being punched into POS again — double-counting risk"


@dataclass
class OrderListingRow:
    order_no: object
    order_type: str
    amount: decimal.Decimal
    discount: decimal.Decimal
    status: str
    created: datetime
    business_date: date


@dataclass
class ParseError:
    line: int
    message: str


def _parse_created(value) -> datetime:
    if hasattr(value, "date") and not isinstance(value, str):
        return value
    return datetime.strptime(str(value).strip(), "%d %b %Y %H:%M:%S")


def parse_order_listing_xlsx(path: str) -> tuple[list[OrderListingRow], list[ParseError]]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx = None
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).strip() == "Order No.":
            header_row_idx = i
            header = row
            break
    if header_row_idx is None:
        raise ValueError("Could not find header row with 'Order No.' in column A")

    col = {name: idx for idx, name in enumerate(header) if name}
    missing = [c for c in REQUIRED_COLUMNS if c not in col]
    if missing:
        raise ValueError(f"Missing expected column(s): {', '.join(missing)}")

    rows: list[OrderListingRow] = []
    errors: list[ParseError] = []

    for line_no, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        order_no = row[col["Order No."]]
        if order_no is None:
            continue
        try:
            order_type = (row[col["Order Type"]] or "").strip()
            amount = decimal.Decimal(str(row[col["My Amount (₹)"]] or "0"))
            discount = decimal.Decimal(str(row[col["Total Discount (₹)"]] or "0"))
            status = (row[col["Status"]] or "").strip()
            created = _parse_created(row[col["Created"]])
            business_date = (created - timedelta(hours=6)).date()

            rows.append(OrderListingRow(
                order_no=order_no, order_type=order_type, amount=amount,
                discount=discount, status=status, created=created,
                business_date=business_date,
            ))
        except Exception as exc:
            errors.append(ParseError(line_no, f"Order No. {order_no!r}: {exc}"))

    return rows, errors


def aggregate_by_day(rows: list[OrderListingRow], upload_date: date) -> dict[date, dict]:
    """Per-business-date {count, amount} for included rows: excludes
    Status='Cancelled' and business_date >= upload_date (today's export is
    incomplete). 'Saved' status is included, matching the item report."""
    totals: dict[date, dict] = {}
    for r in rows:
        if r.status == "Cancelled":
            continue
        if r.business_date >= upload_date:
            continue
        bucket = totals.setdefault(r.business_date, {"count": 0, "amount": decimal.Decimal("0")})
        bucket["count"] += 1
        bucket["amount"] += r.amount
    return totals


def check_aggregator_alarm(rows: list[OrderListingRow]) -> bool:
    """True if any Delivery-type row lands after the date staff stopped
    manually punching aggregator orders into the POS."""
    return any(r.order_type == "Delivery" and r.business_date > ALARM_CUTOFF for r in rows)


def find_unpaired_diffs(daily_amounts: dict[date, decimal.Decimal], tolerance=decimal.Decimal("1")) -> list[tuple[date, decimal.Decimal]]:
    """Days where |diff| > tolerance and it does NOT cancel out with an
    immediately adjacent day (a boundary-attribution artifact of the -6h
    business-date rule shows up as a same-magnitude, opposite-sign diff on
    the day before or after — that pairing is acceptable noise, not a real
    mismatch)."""
    unpaired = []
    for d, diff in daily_amounts.items():
        if abs(diff) <= tolerance:
            continue
        prev_diff = daily_amounts.get(d - timedelta(days=1))
        next_diff = daily_amounts.get(d + timedelta(days=1))
        if prev_diff is not None and abs(prev_diff + diff) <= tolerance:
            continue
        if next_diff is not None and abs(next_diff + diff) <= tolerance:
            continue
        unpaired.append((d, diff))
    return sorted(unpaired)


def upsert_order_counts(db: Session, daily_counts: dict[date, dict]) -> tuple[int, list[date]]:
    """UPDATE daily_channel_sales.orders (channel='petpooja') for each date
    that already has a row — net_sales is owned by the item-report upload,
    so a date with no existing row is skipped rather than inserted with a
    fabricated net_sales. Returns (dates_updated, dates_skipped)."""
    updated = 0
    skipped: list[date] = []
    for business_date, agg in daily_counts.items():
        result = db.query(DailyChannelSales).filter(
            DailyChannelSales.channel == "petpooja",
            DailyChannelSales.business_date == business_date,
        ).update({"orders": agg["count"]}, synchronize_session=False)
        if result:
            updated += 1
        else:
            skipped.append(business_date)
    return updated, skipped


def cross_check_amounts(db: Session, daily_counts: dict[date, dict]) -> dict[date, decimal.Decimal]:
    """{business_date: order_listing_amount - daily_channel_sales.net_sales}
    for dates that have an existing petpooja row (net_sales comes from the
    item report, this file's SUM(My Amount) is the crosscheck)."""
    diffs: dict[date, decimal.Decimal] = {}
    for business_date, agg in daily_counts.items():
        net_sales = db.query(DailyChannelSales.net_sales).filter(
            DailyChannelSales.channel == "petpooja",
            DailyChannelSales.business_date == business_date,
        ).scalar()
        if net_sales is None:
            continue
        diffs[business_date] = agg["amount"] - net_sales
    return diffs
